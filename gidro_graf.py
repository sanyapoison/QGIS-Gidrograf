# -*- coding: utf-8 -*-

# Инициализация библиотек Pyqt и Qgis
from PyQt4 import Qt, QtCore, QtGui, uic
from PyQt4.QtCore import *
from PyQt4.QtGui import *
from PyQt4.QtSql import *

from qgis.utils import *
from qgis.core import *
from qgis.gui import *

# Инициализация дополнительных библиотек
from ConfigParser import *
from openpyxl import *
from math import *

# Инициализация ресурсов
import resources_rc
import datetime
import psycopg2
import os.path
import time
import sip                       
        
class PointTool(QgsMapTool):       
    def __init__(self, canvas, name_slot_coord, name_slot_click, options):
        QgsMapTool.__init__(self, canvas)        
        self.canvas = canvas      
        self.gidrograf = gidro_graf(self)
        self.name_slot_coord = name_slot_coord        
        self.name_slot_click = name_slot_click        
        self.options = options
               
    def canvasMoveEvent(self, event):
        x = event.pos().x()
        y = event.pos().y()
        point = self.canvas.getCoordinateTransform().toMapCoordinates(x, y)
        
        #сигналы координат --------------------------------------------------------
        #---------------------------------------------
        if(self.name_slot_coord == "handle_mouse_coord_ver"):
            self.gidrograf.handle_mouse_coord_ver(point, self.options[3], self.canvas)        
   
        elif(self.name_slot_coord == "def_set_geometry_uzel_coord"):
            self.gidrograf.def_set_geometry_uzel_coord(point, self.options[0], self.options[1], self.canvas) 
   
        elif(self.name_slot_coord == "handle_mouse_coord_ver_for_copy"):
            self.gidrograf.handle_mouse_coord_ver_for_copy(point, self.options[0], self.options[2], self.canvas)            
   
        elif(self.name_slot_coord == "get_coord_for_add_reb"):
            self.gidrograf.get_coord_for_add_reb(point, self.options[0], self.options[1], self.options[2], self.options[3], self.options[4], self.options[5], self.options[6], self.options[7], self.options[8], self.options[9], self.options[10], self.options[11], self.canvas)                           
          
    def canvasPressEvent(self, event):                
        x = event.pos().x()
        y = event.pos().y()
        point = self.canvas.getCoordinateTransform().toMapCoordinates(x, y)

        #сигналы кнопок --------------------------------------------------------
        #---------------------------------------------
        if(self.name_slot_click == "select_uzel_in_canvas"):
            self.gidrograf.select_uzel_in_canvas(point, event.buttons(), self.options[0], self.options[1])
        
        elif(self.name_slot_click == "select_rebro_in_canvas"):
            self.gidrograf.select_rebro_in_canvas(point, event.buttons(), self.options[0], self.options[1])
            
        #---------------------------------------------
        elif(self.name_slot_click == "get_coord_for_add_uzel"):
            self.gidrograf.get_coord_for_add_uzel(point, event.buttons(), self.options[0], self.options[1], self.options[2], self.options[3], self.options[4], self.options[5], self.options[6], self.canvas)          
   
        elif(self.name_slot_click == "get_id_for_edit_uzel"):
            self.gidrograf.get_id_for_edit_uzel(point, event.buttons(), self.options[0], self.options[1], self.options[2], self.options[3], self.canvas)          
   
        elif(self.name_slot_click == "get_id_for_rem_uzel"):
            self.gidrograf.get_id_for_rem_uzel(point, event.buttons(), self.options[0], self.options[1], self.options[2], self.options[3], self.options[4], self.canvas)          
      
        elif(self.name_slot_click == "def_set_geometry_uzel_click"):
            self.gidrograf.def_set_geometry_uzel_click(point, event.buttons(), self.options[0], self.options[1], self.options[2], self.options[4], self.canvas)          
 
        elif(self.name_slot_click == "get_id_for_del_uzel"):
            self.gidrograf.get_id_for_del_uzel(point, event.buttons(), self.options[0], self.options[1], self.options[2], self.canvas)          
 
        elif(self.name_slot_click == "get_id_for_copy_uzel"):
            self.gidrograf.get_id_for_copy_uzel(point, event.buttons(), self.options[0], self.options[1], self.options[2], self.canvas)          
      
        elif(self.name_slot_click == "get_coord_for_add_uzel_for_copy"):
            self.gidrograf.get_coord_for_add_uzel_for_copy(point, event.buttons(), self.options[1], self.canvas)          
            
        #---------------------------------------------
        elif(self.name_slot_click == "get_id_for_add_reb"):
            self.gidrograf.get_id_for_add_reb(point, event.buttons(), self.options[0], self.options[1], self.options[2], self.options[3], self.options[4], self.options[5], self.options[6], self.options[7], self.options[8], self.options[9], self.options[10], self.options[11], self.canvas)                           
 
        elif(self.name_slot_click == "get_id_for_del_rebro"):
            self.gidrograf.get_id_for_del_rebro(point, event.buttons(), self.options[0], self.options[1], self.options[2], self.canvas)   
          
          
          
          
    def canvasReleaseEvent(self, event):         
        if(self.name_slot_click == "get_id_for_add_reb"):
            None
            #self.gidrograf.def_off_signal_for_add_reb(point, event.buttons(), self.options[0], self.options[1], self.options[2], self.options[3], self.options[4], self.options[5], self.options[6], self.options[7], self.options[8], self.options[9], self.options[10], self.options[11], self.canvas)                           
          
        
#главный класс gidro_graf
class gidro_graf:    
    #инициализация компонентов и локали 
    def __init__(self, iface):
        self.iface = iface                        
        self.plugin_dir = os.path.dirname(__file__)
                
        locale = QSettings().value("locale/userLocale")[0:2]
        localePath = os.path.join(self.plugin_dir, 'i18n', 'gidro_graf_{}.qm'.format(locale))

        if os.path.exists(localePath):
            self.translator = QTranslator()
            self.translator.load(localePath)

            if qVersion() > '4.3.3':
                QCoreApplication.installTranslator(self.translator)               
                
    #инициализация графики            
    def initGui(self):   
        self.action_reload_gidrograf = QAction( QIcon(self.plugin_dir + "/icons/reload_gidrograf.png"), u"Перзагрузить ГидроГраф", self.iface.mainWindow())
        self.action_reload_gidrograf.triggered.connect(self.def_reload_gidrograf)
        self.iface.addToolBarIcon(self.action_reload_gidrograf)
    
        #добавление апплета ГидроГраф в менюбар        
        self.add_main_menu_gidrograf()
        #инициализация тулбара
        self.add_toolbar_gidrograf()   

        #переменные для хранение конфигураций
        self.isset_lay_ver = None
        self.isset_nam_ver = None
        self.load_lay_ver  = False 
        self.isset_lay_reb = None
        self.isset_nam_reb = None         
        self.load_lay_reb  = False
    
        #проверка загруженности слоев
        layers = self.iface.legendInterface().layers()
        load_layer_list = []
        for layer in layers:
            layerType = layer.type()
            if(layerType == QgsMapLayer.VectorLayer):
                load_layer_list.append(unicode(layer.name()))        
                
        if(len(load_layer_list)):
            self.read_config_to_load_layers()
            for index_active_layer in self.iface.legendInterface().layers():
                name_active_layer = unicode(index_active_layer.name())
                                    
                if(unicode("zdania_vodokanal_new_2") == name_active_layer):                
                    self.def_load_layer_zdan()
                    self.action_add_layer_zdan.setEnabled(False)

                if(unicode(self.isset_nam_reb) == name_active_layer):                
                    self.def_load_layer_reb()                                           
                    self.action_add_layer_reb.setEnabled(False)                    
                    
                if(unicode(self.isset_nam_ver) == name_active_layer):            
                    self.def_load_layer_ver()
                    self.action_add_layer_ver.setEnabled(False)

                

                
    #выгрузка плагина    
    def unload(self):
        #удаление иконки экшина перезагрузки
        self.iface.removeToolBarIcon(self.action_reload_gidrograf)
        #удаление апплета из менюбара
        self.delete_main_menu_gidrograf()       
        #выгрузка тулбара 
        self.delete_toolbar_gidrograf()      

    def def_reload_gidrograf(self):    
        reloadPlugin('Gidrograf')
    
    def connect_signal_point_tool_init(self, canvas, name_signal_coord, name_signal_click, options):
        canvas.setMapTool(PointTool(canvas, name_signal_coord, name_signal_click, options))
        
    ########################################################################################################################################################
    ##########                    MENUBAR                   ################################################################################################
    ########################################################################################################################################################

    #добавление менюбара
    def add_main_menu_gidrograf(self):                    
        self.menu = QMenu(self.iface.mainWindow())
        self.menu.setObjectName("gidrograf_Menu")
        self.menu.setTitle(u"ГидроГраф")        
        
        self.action_add_layer_zdan = QAction(QIcon(self.plugin_dir + "/icons/icon.png"), u"Загрузить слои зданий и дорог", self.iface.mainWindow())
        self.action_add_layer_ver = QAction(QIcon(self.plugin_dir + "/icons/vershini_menu.png"), u"Загрузить слой узлов", self.iface.mainWindow())
        self.action_add_layer_reb = QAction(QIcon(self.plugin_dir + "/icons/rebra_menu.png"), u"Загрузить слой ребер", self.iface.mainWindow())        
                
        # функционал кнопок
        QObject.connect(self.action_add_layer_zdan, SIGNAL("triggered()"), self.def_load_layer_zdan)
        QObject.connect(self.action_add_layer_ver, SIGNAL("triggered()"), self.def_load_layer_ver)
        QObject.connect(self.action_add_layer_reb, SIGNAL("triggered()"), self.def_load_layer_reb)
        
        
        self.action_add_layer_reb.setEnabled(False)   
        self.action_add_layer_ver.setEnabled(False) 
        
        self.menu.addAction(self.action_add_layer_zdan)
        self.menu.addAction(self.action_add_layer_ver)
        self.menu.addAction(self.action_add_layer_reb)
        
        menuBar = self.iface.mainWindow().menuBar()
        menuBar.insertMenu(self.iface.firstRightStandardMenu().menuAction(), self.menu)                        
                        
    #удаление апплета из менюбара    
    def delete_main_menu_gidrograf(self):  
        self.menu.deleteLater()                     
        
    ##############################################
    ##########         Функиции         ##########
    ##############################################
    
    #чтение конфигураций для загрузки слоев    
    def read_config_to_load_layers(self):   
        if(not(self.isset_lay_ver) or not(self.isset_nam_ver) or not(self.isset_nam_reb) or not(self.isset_nam_reb)):            
            try:
                config = RawConfigParser()
                file_cfg = self.plugin_dir + "/config/config.ini"
                config.read(file_cfg)

                self.isset_lay_ver = config.getboolean("db_conf", "bool_layer_ver")
                self.isset_nam_ver = config.get("db_conf", "name_layer_ver")
                self.isset_lay_reb = config.getboolean("db_conf", "bool_layer_reb")
                self.isset_nam_reb = config.get("db_conf", "name_layer_reb")            
                    
            except:
                QtGui.QMessageBox.warning(self.iface.mainWindow(), u"Ошибки файла конфигураций", u"Не удалось прочитать файл конфигураций")     

    # получение списка слоев в базе            
    def get_list_layers_in_bd(self):        
        self.connect_bd = False
        try:
            conn = psycopg2.connect("dbname='vodokanal_data' user='postgres' host='localhost' password='postgres'")                
            cur_con = conn.cursor()
            cur_con.execute("SELECT table_name FROM information_schema.tables WHERE table_schema = 'public'")
            numrows_con = int(cur_con.rowcount)
                    
            self.layers_list = []                       
            for x in range(numrows_con):
                memory_con = cur_con.fetchone()
                self.layers_list.append(memory_con[0])                  
                  
            cur_con.close()
            conn.close()                 
                
            self.connect_bd = True
        except:
            QtGui.QMessageBox.warning(self.iface.mainWindow(), u"Соединение с БД", u"Не удалось соединиться с БД. \nНе удалось получить список слоев в БД")
        
    #загрузка слоя узлов из базы   
    def def_load_layer_ver(self):        
        self.read_config_to_load_layers()
        self.get_list_layers_in_bd()
        self.load_lay_ver = True
        
        if(self.connect_bd): 
            #список активных слоев
            layers = self.iface.legendInterface().layers()
            load_layer_list = []
            for layer in layers:
                layerType = layer.type()
                if(layerType == QgsMapLayer.VectorLayer):
                    load_layer_list.append(unicode(layer.name()))         
        
            #загрузка слоев из базы        
            if(self.isset_nam_ver in self.layers_list):
                if(not(self.isset_nam_ver in load_layer_list)):
                    uri = QgsDataSourceURI()
                    uri.setConnection("localhost", "5432", "vodokanal_data", "postgres", "postgres")
                    uri.setDataSource("public", self.isset_nam_ver, "geom")
                    self.point_layer = self.iface.addVectorLayer(uri.uri(), self.isset_nam_ver, "postgres")
                        
                else:
                    for layer in layers:
                        layerType = layer.type()
                        if((layerType == QgsMapLayer.VectorLayer) and (layer.name() == self.isset_nam_ver)):
                            self.point_layer = layer
                            self.iface.mapCanvas().refresh()
                            if hasattr(self.point_layer, "setCacheImage"):
                                self.point_layer.setCacheImage(None)
                            self.point_layer.triggerRepaint() 
                            break
                                            
                #отключить событие загрузки слоя вершин                            
                self.action_add_layer_ver.setEnabled(False)   
                self.action_add_layer_reb.setEnabled(False)
                
                #активация кнопок тулбара
                self.panel_instrumentov.btn_regim_uzlov.setEnabled(True)
                self.panel_instrumentov.btn_spisok_videleniv.setEnabled(True)
                self.panel_instrumentov.btn_poisk_objectov_sloev.setEnabled(True)
                self.panel_instrumentov.btn_support_create.setEnabled(True)                            
                self.panel_instrumentov.btn_create_uzel.setEnabled(True)                            
                self.panel_instrumentov.btn_edt_uzel.setEnabled(True)                            
                self.panel_instrumentov.btn_rem_uzel.setEnabled(True)                            
                self.panel_instrumentov.btn_copy_uzel.setEnabled(True)                            
                self.panel_instrumentov.btn_del_uzel.setEnabled(True)                            
                
            else:
                QtGui.QMessageBox.warning(self.iface.mainWindow(), u"Слой не найден", u"Указаный слой вершин отсутствует в базе. \nСоздайте слой вершин!")                                
        else:
            QtGui.QMessageBox.warning(self.iface.mainWindow(), u"Соединение с БД", u"Не установлено соединение с БД!")                                               
                                               
    #загрузка слоя ребер из базы           
    def def_load_layer_reb(self):
        self.read_config_to_load_layers()
        self.get_list_layers_in_bd()
        self.load_lay_reb = True
        if(self.connect_bd): 
            #список активных слоев
            layers = self.iface.legendInterface().layers()
            load_layer_list = []
            for layer in layers:
                layerType = layer.type()
                if(layerType == QgsMapLayer.VectorLayer):
                    load_layer_list.append(unicode(layer.name()))         
        
            #загрузка слоев из базы        
            if(self.isset_nam_reb in self.layers_list):
                if(not(self.isset_nam_reb in load_layer_list)):
                    uri = QgsDataSourceURI()
                    uri.setConnection("localhost", "5432", "vodokanal_data", "postgres", "postgres")
                    uri.setDataSource("public", self.isset_nam_reb, "geom")
                    self.line_layer = self.iface.addVectorLayer(uri.uri(), self.isset_nam_reb, "postgres")
                        
                else:
                    for layer in layers:
                        layerType = layer.type()
                        if((layerType == QgsMapLayer.VectorLayer) and (layer.name() == self.isset_nam_reb)):
                            self.line_layer = layer
                            self.iface.mapCanvas().refresh()
                            if hasattr(self.line_layer, "setCacheImage"):
                                self.line_layer.setCacheImage(None)
                            self.line_layer.triggerRepaint() 
                            break
                                            
                #отключить событие загрузки слоя ребер                            
                self.action_add_layer_reb.setEnabled(False)   
                self.action_add_layer_ver.setEnabled(True)   

                #активация кнопок тулбара
                self.panel_instrumentov.btn_regim_reber.setEnabled(True)
                self.panel_instrumentov.btn_spisok_videleniv.setEnabled(True)
                self.panel_instrumentov.btn_poisk_objectov_sloev.setEnabled(True)
                self.panel_instrumentov.btn_support_create.setEnabled(True)
                self.panel_instrumentov.btn_create_reb.setEnabled(True)                                                        
                self.panel_instrumentov.btn_del_reb.setEnabled(True)
                                            
            else:
                QtGui.QMessageBox.information(self.iface.mainWindow(), u"Слой не найден", u"Указаный слой ребер отсутствует в базе. \nСоздайте слой ребер!")                                
        else:
            QtGui.QMessageBox.warning(self.iface.mainWindow(), u"Соединение с БД", u"Не установлено соединение с БД!")

    #загрузка слоя ребер из базы           
    def def_load_layer_zdan(self):
        self.get_list_layers_in_bd()
        if(self.connect_bd): 
            #список активных слоев
            layers = self.iface.legendInterface().layers()
            load_layer_list = []
            for layer in layers:
                layerType = layer.type()
                if(layerType == QgsMapLayer.VectorLayer):
                    load_layer_list.append(unicode(layer.name()))         
        
            #загрузка слоев из базы        
            if(unicode("zdania_vodokanal_new_2") in self.layers_list):
                if(not(unicode("zdania_vodokanal_new_2") in load_layer_list)):
                    uri = QgsDataSourceURI()
                    uri.setConnection("localhost", "5432", "vodokanal_data", "postgres", "postgres")
                    uri.setDataSource("public", unicode("zdania_vodokanal_new_2"), "geom")
                    self.zdan_layer = self.iface.addVectorLayer(uri.uri(), unicode("zdania_vodokanal_new_2"), "postgres")
                        
                else:
                    for layer in layers:
                        layerType = layer.type()
                        if((layerType == QgsMapLayer.VectorLayer) and (layer.name() == unicode("zdania_vodokanal_new_2"))):
                            self.zdan_layer = layer
                            self.iface.mapCanvas().refresh()
                            if hasattr(self.zdan_layer, "setCacheImage"):
                                self.zdan_layer.setCacheImage(None)
                            self.zdan_layer.triggerRepaint() 
                            break
                                            
                #отключить событие загрузки слоя ребер                            
                self.action_add_layer_zdan.setEnabled(False)
                self.action_add_layer_reb.setEnabled(True)
 
           #загрузка слоев из базы        
            if(unicode("open_street_map") in self.layers_list):
                if(not(unicode("open_street_map") in load_layer_list)):
                    uri = QgsDataSourceURI()
                    uri.setConnection("localhost", "5432", "vodokanal_data", "postgres", "postgres")
                    uri.setDataSource("public", unicode("open_street_map"), "geom")
                    self.dorog_layer = self.iface.addVectorLayer(uri.uri(), unicode("open_street_map"), "postgres")
                        
                else:
                    for layer in layers:
                        layerType = layer.type()
                        if((layerType == QgsMapLayer.VectorLayer) and (layer.name() == unicode("open_street_map"))):
                            self.dorog_layer = layer
                            self.iface.mapCanvas().refresh()
                            if hasattr(self.dorog_layer, "setCacheImage"):
                                self.dorog_layer.setCacheImage(None)
                            self.dorog_layer.triggerRepaint() 
                            break
                                            
                #отключить событие загрузки слоя ребер                            
                self.action_add_layer_zdan.setEnabled(False)
                self.action_add_layer_reb.setEnabled(True)
  
                                            
            else:
                QtGui.QMessageBox.information(self.iface.mainWindow(), u"Слой не найден", u"Указаный слой ребер отсутствует в базе. \nСоздайте слой ребер!")                                
        else:
            QtGui.QMessageBox.warning(self.iface.mainWindow(), u"Соединение с БД", u"Не установлено соединение с БД!") 
 
    ########################################################################################################################################################
    ##########                    TOOLBAR                   ################################################################################################
    ########################################################################################################################################################
    def add_toolbar_gidrograf(self):     
        self.panel_instrumentov = uic.loadUi(self.plugin_dir + "/forms/panel_instrumentov.ui")
        self.apdockwidget_panel_instrumentov = QDockWidget(u"" , self.iface.mainWindow())
        self.apdockwidget_panel_instrumentov.setObjectName(u"Гидрограф")
        self.apdockwidget_panel_instrumentov.setWidget(self.panel_instrumentov)
        self.iface.addDockWidget( Qt.RightDockWidgetArea, self.apdockwidget_panel_instrumentov)
            
        # функционал кнопок
        QtCore.QObject.connect(self.panel_instrumentov.btn_regim_uzlov,          QtCore.SIGNAL("clicked()"), self.init_regim_uzlov)
        QtCore.QObject.connect(self.panel_instrumentov.btn_regim_reber,          QtCore.SIGNAL("clicked()"), self.init_regim_reber)        
        
        QtCore.QObject.connect(self.panel_instrumentov.btn_spisok_videleniv,     QtCore.SIGNAL("clicked()"), self.get_list_selected_object)       
        QtCore.QObject.connect(self.panel_instrumentov.btn_poisk_objectov_sloev, QtCore.SIGNAL("clicked()"), self.get_form_search)
        
        QtCore.QObject.connect(self.panel_instrumentov.btn_support_create,       QtCore.SIGNAL("clicked()"), self.get_form_support)
        
        QtCore.QObject.connect(self.panel_instrumentov.btn_create_uzel,          QtCore.SIGNAL("clicked()"), self.def_add_uzel)
        QtCore.QObject.connect(self.panel_instrumentov.btn_edt_uzel,             QtCore.SIGNAL("clicked()"), self.def_edit_uzel)
        QtCore.QObject.connect(self.panel_instrumentov.btn_copy_uzel,             QtCore.SIGNAL("clicked()"), self.def_copy_uzel)
        QtCore.QObject.connect(self.panel_instrumentov.btn_rem_uzel,             QtCore.SIGNAL("clicked()"), self.def_remont_uzla)
        QtCore.QObject.connect(self.panel_instrumentov.btn_del_uzel,             QtCore.SIGNAL("clicked()"), self.def_del_uzel)

        QtCore.QObject.connect(self.panel_instrumentov.btn_create_reb,          QtCore.SIGNAL("clicked()"), self.def_add_reb)
        QtCore.QObject.connect(self.panel_instrumentov.btn_del_reb,             QtCore.SIGNAL("clicked()"), self.def_del_reb)

        
        self.connect_signal_point_tool_init(self.iface.mapCanvas(), "", "", [])
    #выгрузка тулбара     
    def delete_toolbar_gidrograf(self):     
        try:                 
            self.iface.removeDockWidget(self.apdockwidget_panel_instrumentov)
            sip.delete(self.apdockwidget_panel_instrumentov)                       
        except:
            QtGui.QMessageBox.warning(self.iface.mainWindow(), u"Ошибка закрытия", u"Не удалось корректно закрыть редактор")    
        
    #инициализация флага обработки узлов    
    def init_regim_uzlov(self):                 
        if(self.panel_instrumentov.btn_regim_reber.isChecked()):
            self.panel_instrumentov.btn_regim_reber.setChecked(False)
                            
        if(self.panel_instrumentov.btn_regim_uzlov.isChecked()):    
            self.flag_perekluchenia_uzlov_vershin = u"init_regim_uzlov"
        else:
            self.flag_perekluchenia_uzlov_vershin = u"init_none"
        
        self.connect_signal_point_tool_init(self.iface.mapCanvas(), "", "select_uzel_in_canvas", [self.point_layer, self.panel_instrumentov])
        
    #выделение вершины           
    def select_uzel_in_canvas(self, point, button, point_layer, panel_instrumentov):
        if(panel_instrumentov.btn_regim_uzlov.isChecked()):
            #определение id ближайшего элемента к курсору
            provider = point_layer.dataProvider()                 
            spIndex = QgsSpatialIndex()
            feat = QgsFeature()
            fit = provider.getFeatures() 
            while fit.nextFeature(feat):
                spIndex.insertFeature(feat)
            pt = QgsPoint(point.x(), point.y())			
            nearestIds = spIndex.nearestNeighbor(pt,1)            
            id_feature = nearestIds[0]
            
            # если нажата левая кнопка мыши выделить объект
            if(button == Qt.LeftButton):    
                point_layer.select(id_feature)                   
                             
            # если нажата правая кнопка мыши отменить выделение                             
            if(button == Qt.RightButton):            
                point_layer.deselect(id_feature)    
       
    #инициализация обработки флага ребер
    def init_regim_reber(self):        
        if(self.panel_instrumentov.btn_regim_uzlov.isChecked()):
            self.panel_instrumentov.btn_regim_uzlov.setChecked(False) 
                
        if(self.panel_instrumentov.btn_regim_reber.isChecked()):
            self.flag_perekluchenia_uzlov_vershin = u"init_regim_reber"                    
        else:    
            self.flag_perekluchenia_uzlov_vershin = u"init_none"                                         
        
        #сигнал выбора ребра по id
        self.connect_signal_point_tool_init(self.iface.mapCanvas(), "", "select_rebro_in_canvas", [self.line_layer, self.panel_instrumentov])                
        
    #выделение ребра           
    def select_rebro_in_canvas(self, point, button, line_layer, panel_instrumentov):
        if(panel_instrumentov.btn_regim_reber.isChecked()):
            #определение id ближайшего элемента к курсору
            provider = line_layer.dataProvider()                 
            spIndex = QgsSpatialIndex()
            feat = QgsFeature()
            fit = provider.getFeatures() 
            while fit.nextFeature(feat):
                spIndex.insertFeature(feat)
            pt = QgsPoint(point.x(), point.y())			
            nearestIds = spIndex.nearestNeighbor(pt,1)            
            id_feature = nearestIds[0]
            
            # если нажата левая кнопка мыши выделить объект
            if(button == Qt.LeftButton):    
                line_layer.select(id_feature)                   
                  
            # если нажата правая кнопка мыши отменить выделение
            if(button == Qt.RightButton):            
                line_layer.deselect(id_feature)         
          
    #форма списков выделенных объектов     
    def get_list_selected_object(self):        
        file_ui = self.plugin_dir + "/forms/forma_list_selected.ui"
        self.forma_list_selected = uic.loadUi(file_ui)
        self.forma_list_selected.setWindowTitle(u'Список выделенных объектов')        
        self.forma_list_selected.setWindowModality(QtCore.Qt.ApplicationModal)             
                
        QtCore.QObject.connect(self.forma_list_selected.btn_cls_form_list_selected, QtCore.SIGNAL("clicked()"), self.def_close_form_selected)        
        
        if(self.load_lay_ver):
            QtCore.QObject.connect(self.forma_list_selected.btn_disselected_uzla, QtCore.SIGNAL("clicked()"), self.def_disselected_uzel)        
            QtCore.QObject.connect(self.forma_list_selected.btn_disselected_uzlov, QtCore.SIGNAL("clicked()"), self.def_disselected_vseh_uzlov)        
            QtCore.QObject.connect(self.forma_list_selected.btn_selected_all_uzlov, QtCore.SIGNAL("clicked()"), self.def_selected_vseh_uzlov)        

            self.forma_list_selected.btn_selected_all_uzlov.setEnabled(True)        
            count_selected_points = self.point_layer.selectedFeatureCount()        
            if(count_selected_points > 0):
                self.forma_list_selected.btn_disselected_uzla.setEnabled(True)    
                self.forma_list_selected.btn_disselected_uzlov.setEnabled(True)                                           
                get_list_ids_selected_points = self.point_layer.selectedFeatures()                    
                for selected_point in get_list_ids_selected_points:
                    atributes_select_point = selected_point.attributes()  
                    item = unicode(selected_point.id()) + u" | " + unicode(atributes_select_point[2]) + u" | " + unicode(atributes_select_point[1])                
                    self.forma_list_selected.table_selected_uzlov.addItem(QtGui.QListWidgetItem(item))

        else:
            self.forma_list_selected.btn_disselected_uzla.setEnabled(False)
            self.forma_list_selected.btn_disselected_uzlov.setEnabled(False)
            self.forma_list_selected.btn_selected_all_uzlov.setEnabled(False)        
                    
                    
        if(self.load_lay_reb):                    
            try:
                QtCore.QObject.connect(self.forma_list_selected.btn_disselected_rebra, QtCore.SIGNAL("clicked()"), self.def_disselected_rebra)        
                QtCore.QObject.connect(self.forma_list_selected.btn_disselected_reber, QtCore.SIGNAL("clicked()"), self.def_disselected_vseh_reber)        
                QtCore.QObject.connect(self.forma_list_selected.btn_selected_all_reber, QtCore.SIGNAL("clicked()"), self.def_selected_vseh_reber)        
                    
                self.forma_list_selected.btn_selected_all_reber.setEnabled(True)                
                count_selected_lines = self.line_layer.selectedFeatureCount()
                if(count_selected_lines > 0):
                    self.forma_list_selected.btn_disselected_rebra.setEnabled(True)    
                    self.forma_list_selected.btn_disselected_reber.setEnabled(True)                                                                                    
                    get_list_ids_selected_lines = self.line_layer.selectedFeatures()                
                    for selected_line in get_list_ids_selected_lines:
                        atributes_select_line = selected_line.attributes()  
                        item = unicode(selected_line.id()) + u" | " + unicode(atributes_select_line[4])                
                        self.forma_list_selected.table_selected_reber.addItem(QtGui.QListWidgetItem(item))
            except:
                print "error: get_list_selected_object"
        else:
            self.forma_list_selected.btn_disselected_rebra.setEnabled(False)
            self.forma_list_selected.btn_disselected_reber.setEnabled(False)
            self.forma_list_selected.btn_selected_all_reber.setEnabled(False)        

                    
        self.forma_list_selected.show()
     
    #отмена выделения узла
    def def_disselected_uzel(self):
        index_selected_item = self.forma_list_selected.table_selected_uzlov.currentRow()

        if(index_selected_item >= 0):            
            try:
                text_selected_item = unicode(self.forma_list_selected.table_selected_uzlov.item(index_selected_item).text())
                current_item_for_disselected = True
            except:
                QtGui.QMessageBox.warning(self.iface.mainWindow(), u"Ошибка отмены выделения", u"Не удалось выполнить отмену выделения")    
                current_item_for_disselected = False
             
            if(current_item_for_disselected):
                id_disselected_feature = int(text_selected_item.split(u" | ")[0])
                self.point_layer.deselect(id_disselected_feature)
                self.forma_list_selected.table_selected_uzlov.takeItem(index_selected_item)
                
                count_selected_points = self.point_layer.selectedFeatureCount()
                if(count_selected_points == 0):
                    self.forma_list_selected.btn_disselected_uzla.setEnabled(False)    
                    
                if(count_selected_points == 1):    
                    self.forma_list_selected.btn_disselected_uzlov.setEnabled(False)
            
    #отмена выделения всех узлов
    def def_disselected_vseh_uzlov(self):    
        self.forma_list_selected.table_selected_uzlov.clear()
        self.point_layer.removeSelection()
        self.forma_list_selected.btn_disselected_uzla.setEnabled(False)               
        self.forma_list_selected.btn_disselected_uzlov.setEnabled(False)                

    #выделить все узлы    
    def def_selected_vseh_uzlov(self):
        self.point_layer.selectAll()
        count_selected_points = self.point_layer.selectedFeatureCount()        
        self.forma_list_selected.btn_selected_all_uzlov.setEnabled(True)        
        if(count_selected_points > 0):
            self.forma_list_selected.btn_disselected_uzla.setEnabled(True)    
            self.forma_list_selected.btn_disselected_uzlov.setEnabled(True)                                           
            get_list_ids_selected_points = self.point_layer.selectedFeatures()                    
            for selected_point in get_list_ids_selected_points:
                atributes_select_point = selected_point.attributes()  
                item = unicode(selected_point.id()) + u" | " + unicode(atributes_select_point[2]) + u" | " + unicode(atributes_select_point[1])                
                self.forma_list_selected.table_selected_uzlov.addItem(QtGui.QListWidgetItem(item))
                
    #отмена выделения ребра
    def def_disselected_rebra(self):
        index_selected_item = self.forma_list_selected.table_selected_uzlov.currentRow()
        
        if(index_selected_item >= 0 ):
            try:
                text_selected_item = unicode(self.forma_list_selected.table_selected_reber.item(index_selected_item).text())
                current_item_for_disselected = True
            except:
                QtGui.QMessageBox.warning(self.iface.mainWindow(), u"Ошибка отмены выделения", u"Не удалось выполнить отмену выделения")    
                current_item_for_disselected = False
             
            if(current_item_for_disselected):        
            
                text_selected_item = unicode(self.forma_list_selected.table_selected_reber.item(index_selected_item).text())
                id_disselected_feature = int(text_selected_item.split(u" | ")[0])
                self.line_layer.deselect(id_disselected_feature)
                self.forma_list_selected.table_selected_reber.takeItem(index_selected_item)
                
                count_selected_lines = self.line_layer.selectedFeatureCount()
                if(count_selected_lines == 0):
                    self.forma_list_selected.btn_disselected_rebra.setEnabled(False)    
                    
                if(count_selected_lines == 1):    
                    self.forma_list_selected.btn_disselected_reber.setEnabled(False)
                               
    #отмена выделения всех ребер
    def def_disselected_vseh_reber(self):    
        self.forma_list_selected.table_selected_reber.clear()
        self.line_layer.removeSelection()
        self.forma_list_selected.btn_disselected_rebra.setEnabled(False)               
        self.forma_list_selected.btn_disselected_reber.setEnabled(False)
       
    #выделить все ребра
    def def_selected_vseh_reber(self): 
        self.line_layer.selectAll()    
        self.forma_list_selected.btn_selected_all_reber.setEnabled(True)                
        count_selected_lines = self.line_layer.selectedFeatureCount()
        if(count_selected_lines > 0):
            self.forma_list_selected.btn_disselected_rebra.setEnabled(True)    
            self.forma_list_selected.btn_disselected_reber.setEnabled(True)                                                                                    
            get_list_ids_selected_lines = self.line_layer.selectedFeatures()                
            for selected_line in get_list_ids_selected_lines:
                atributes_select_line = selected_line.attributes()  
                item = unicode(selected_line.id()) + u" | " + unicode(atributes_select_line[4]) + u" | " + unicode(atributes_select_line[1])                
                self.forma_list_selected.table_selected_reber.addItem(QtGui.QListWidgetItem(item))
                
    #закрыть форму выделенных объектов            
    def def_close_form_selected(self):            
        self.forma_list_selected.hide()
          
    #форма поиска объекта     
    def get_form_search(self):
        #инициализация графической формы        
        file_ui = self.plugin_dir + "/forms/forma_poiska.ui"
        self.forma_poiska = uic.loadUi(file_ui)
        self.forma_poiska.setWindowTitle(u'Поиск узлов')        
        self.forma_poiska.setWindowModality(QtCore.Qt.ApplicationModal)  

        QtCore.QObject.connect(self.forma_poiska.btn_close_search, QtCore.SIGNAL("clicked()"), self.def_close_form_poisk)        
        
        if(self.load_lay_ver):
            QtCore.QObject.connect(self.forma_poiska.start_search, QtCore.SIGNAL("clicked()"), self.def_start_search)        
            QtCore.QObject.connect(self.forma_poiska.btn_cancel_search, QtCore.SIGNAL("clicked()"), self.def_cancel_search)        
        else:
            self.forma_poiska.start_search.setEnabled(False)
            self.forma_poiska.btn_cancel_search.setEnabled(False)
            
        self.forma_poiska.show()
                
    #начать поиск объекта
    def def_start_search(self):
        type_search = self.forma_poiska.panel_poiska.currentIndex()                
        
        if(type_search == 0):
            self.forma_poiska.lbl_count_search_elements.setText(u"")
            feature_search_id = unicode(self.forma_poiska.edt_id_feature.text())
            if(feature_search_id != u"" and feature_search_id.isdigit()):
                flag_search_complete = False
                iter = self.point_layer.getFeatures()
                feature_search_id = int(feature_search_id)
                
                for feature in iter:
                    feature_point_id = feature.id()                     
                    if(feature_point_id == feature_search_id):
                        self.feature_point_search_by_id = feature
                        flag_search_complete = True
                        break
                
                if(flag_search_complete):
                    self.backup_search_by_id = True
                    self.backup_feature_points_search_by_id = self.point_layer.selectedFeatures()
                    self.forma_poiska.list_result_search.clear()
                    self.point_layer.removeSelection()
                    
                    self.forma_poiska.btn_cancel_search.setEnabled(True)
                    
                    self.point_layer.select(feature_search_id)
                    atributes_feature_id_search = self.feature_point_search_by_id.attributes()  
                    
                    item = unicode(self.feature_point_search_by_id.id()) + u" | " + unicode(atributes_feature_id_search[2]) + u" | " + unicode(atributes_feature_id_search[1])                
                    self.forma_poiska.list_result_search.addItem(QtGui.QListWidgetItem(item))
                else:
                    QtGui.QMessageBox.information(self.iface.mainWindow(), u"Результат поиска по ID", u"Поиск не дал результатов")                           
            
            else:
                QtGui.QMessageBox.warning(self.iface.mainWindow(), u"Ошибка ввода", u"Введите корректный ID объекта")    
        
        if(type_search == 1):
            feature_search_index_type = self.forma_poiska.cmb_tyfe_feature.currentIndex()
            
            if(feature_search_index_type == 0):
                feature_search_type = u"Колодец"
            if(feature_search_index_type == 1):
                feature_search_type = u"Потребитель"
            if(feature_search_index_type == 2):
                feature_search_type = u"Насосная станция"
            if(feature_search_index_type == 3):
                feature_search_type = u"Заглушка"
            if(feature_search_index_type == 4):
                feature_search_type = u"Источник"
            if(feature_search_index_type == 5):
                feature_search_type = u"Канализация"
            if(feature_search_index_type == 6):
                feature_search_type = u"Развилка"
                
            flag_search_complete = False 
            iter = self.point_layer.getFeatures()            
            list_search_selected_by_type = []
            
            for feature in iter:
                feature_point_id = feature.id()                     
                if(unicode(feature.attributes()[2]) == feature_search_type):
                    list_search_selected_by_type.append(feature)
                                                          
            if(len(list_search_selected_by_type) > 0):
                self.backup_search_by_type = True
                self.backup_feature_points_search_by_type = self.point_layer.selectedFeatures()
                self.forma_poiska.list_result_search.clear()
                self.point_layer.removeSelection()
                self.forma_poiska.btn_cancel_search.setEnabled(True)
                text_count_features_search_type = u"Найдено объектов: " + unicode(len(list_search_selected_by_type))
                self.forma_poiska.lbl_count_search_elements.setText(text_count_features_search_type)
                
                for feature_search_by_type in list_search_selected_by_type:                    
                    self.point_layer.select(feature_search_by_type.id())
                    atributes_feature_search_by_type = feature_search_by_type.attributes()  
                    
                    item = unicode(feature_search_by_type.id()) + u" | " + unicode(atributes_feature_search_by_type[2]) + u" | " + unicode(atributes_feature_search_by_type[1])                
                    self.forma_poiska.list_result_search.addItem(QtGui.QListWidgetItem(item))                   
            else:
                QtGui.QMessageBox.information(self.iface.mainWindow(), u"Результат поиска по типу", u"Поиск по типу не дал результатов")                           
                       
        if(type_search == 2):
            feature_search_name = unicode(self.forma_poiska.edt_name_feature.text())
            if(feature_search_name != u""):
                iter = self.point_layer.getFeatures()
                list_search_selected_by_name = []
                 
                for feature in iter:
                    feature_point_name = unicode(feature.attributes()[1])                    
                    if(feature_search_name in feature_point_name):
                        list_search_selected_by_name.append(feature)
                              
                if(len(list_search_selected_by_name) > 0):
                    self.backup_search_by_name = True
                    self.backup_feature_points_search_by_name = self.point_layer.selectedFeatures()
                    self.forma_poiska.list_result_search.clear()
                    self.point_layer.removeSelection()
                    self.forma_poiska.btn_cancel_search.setEnabled(True)
                    text_count_features_search_name = u"Найдено объектов: " + unicode(len(list_search_selected_by_name))
                    self.forma_poiska.lbl_count_search_elements.setText(text_count_features_search_name)
                    
                    for feature_search_by_name in list_search_selected_by_name:                    
                        self.point_layer.select(feature_search_by_name.id())
                        atributes_feature_search_by_name = feature_search_by_name.attributes()  
                        
                        item = unicode(feature_search_by_name.id()) + u" | " + unicode(atributes_feature_search_by_name[2]) + u" | " + unicode(atributes_feature_search_by_name[1])                
                        self.forma_poiska.list_result_search.addItem(QtGui.QListWidgetItem(item))                   
                else:
                    QtGui.QMessageBox.information(self.iface.mainWindow(), u"Результат поиска по имени", u"Поиск по имени не дал результатов")               
            else:
                QtGui.QMessageBox.warning(self.iface.mainWindow(), u"Ошибка ввода", u"Введите корректное имя объекта")  
     
    #отменить текущее выделение
    def def_cancel_search(self):
        try:
            if(self.backup_search_by_id):
                self.backup_search_by_id = False
                self.point_layer.removeSelection()     
                self.forma_poiska.edt_id_feature.setText(u"")
                self.forma_poiska.list_result_search.clear()             
                  
                for backup_feature_point in self.backup_feature_points_search_by_id:
                    self.point_layer.select(backup_feature_point.id())                
                    atributes_backup_feature_point = backup_feature_point.attributes() 
                    item = unicode(backup_feature_point.id()) + u" | " + unicode(atributes_backup_feature_point[2]) + u" | " + unicode(atributes_backup_feature_point[1])                
                    self.forma_poiska.list_result_search.addItem(QtGui.QListWidgetItem(item))
                    
            self.forma_poiska.btn_cancel_search.setEnabled(False)
        except:
            None
            
        try:
            if(self.backup_search_by_type):
                self.backup_search_by_type = False
                self.point_layer.removeSelection()     
                self.forma_poiska.lbl_count_search_elements.setText(u"")
                self.forma_poiska.list_result_search.clear()             
                  
                for backup_feature_point in self.backup_feature_points_search_by_type:
                    self.point_layer.select(backup_feature_point.id())                
                    atributes_backup_feature_point = backup_feature_point.attributes() 
                    item = unicode(backup_feature_point.id()) + u" | " + unicode(atributes_backup_feature_point[2]) + u" | " + unicode(atributes_backup_feature_point[1])                
                    self.forma_poiska.list_result_search.addItem(QtGui.QListWidgetItem(item))
                    
            self.forma_poiska.btn_cancel_search.setEnabled(False)  
        except:
            None
            
        try:
            if(self.backup_search_by_name):
                self.backup_search_by_name = False
                self.point_layer.removeSelection()     
                self.forma_poiska.lbl_count_search_elements.setText(u"")
                self.forma_poiska.edt_name_feature.setText(u"")
                self.forma_poiska.list_result_search.clear()             
                  
                for backup_feature_point in self.backup_feature_points_search_by_name:
                    self.point_layer.select(backup_feature_point.id())                
                    atributes_backup_feature_point = backup_feature_point.attributes() 
                    item = unicode(backup_feature_point.id()) + u" | " + unicode(atributes_backup_feature_point[2]) + u" | " + unicode(atributes_backup_feature_point[1])                
                    self.forma_poiska.list_result_search.addItem(QtGui.QListWidgetItem(item))
                    
            self.forma_poiska.btn_cancel_search.setEnabled(False)  
        except:
            None

    #закрыть форму поиска        
    def def_close_form_poisk(self):
        self.forma_poiska.hide()
            
    #сформировать отчет
    def get_form_support(self):  
        file_ui = self.plugin_dir + "/forms/forma_support.ui"
        self.forma_create_support = uic.loadUi(file_ui)
        self.forma_create_support.setWindowTitle(u'Формировать отчет')        
        self.forma_create_support.setWindowModality(QtCore.Qt.ApplicationModal)    
         
        QtCore.QObject.connect(self.forma_create_support.btn_create_support, QtCore.SIGNAL("clicked()"), self.def_create_support)
        QtCore.QObject.connect(self.forma_create_support.btn_close_support, QtCore.SIGNAL("clicked()"), self.def_close_support)

        self.forma_create_support.show()

    #сохранение отчета в файл
    def def_create_support(self):
        index_type_support_item = self.forma_create_support.lst_type_support.currentRow()
        
        if(index_type_support_item >= 0):
            try:
                text_type_support_item = unicode(self.forma_create_support.lst_type_support.item(index_type_support_item).text())
                current_item_for_support = True
            except:
                QtGui.QMessageBox.warning(self.iface.mainWindow(), u"Ошибка типа отчета", u"Не удалось определить тип отчета")    
                current_item_for_support = False
             
            if(current_item_for_support):
                self.forma_create_support.hide()
                filename = unicode(QFileDialog.getSaveFileName(None, u"Сохранение отчета", self.plugin_dir,u"Отчет ГидроГраф (*.xlsx)"))
                dt = datetime.datetime.now()
                date_and_time_for_support = dt.strftime('%Y-%m-%d %H:%M:%S')                
                
                if(text_type_support_item == u"Отчет по узлам" and filename != u""):                                                                  
                    Workbook().save(filename)    
                    workbook = load_workbook(filename)
                    ws = workbook.active
                                        
                    ws['B2'] = u"Отчет по узлам"
                    ws['A3'] = u"Дата и время создания отчета"
                    ws['B3'] = date_and_time_for_support                    
                    ws['A4'] = u"Количество узлов"                   
                    ws['B4'] = unicode(self.point_layer.featureCount())
                    ws['A5'] = u"Количество выделенных узлов"
                    ws['B5'] = unicode(self.point_layer.selectedFeatureCount())                             
                          
                    workbook.save(filename)                    

                if(text_type_support_item == u"Отчет по ребрам" and filename != u""):                                                                  
                    Workbook().save(filename)    
                    workbook = load_workbook(filename)
                    ws = workbook.active
                                        
                    ws['B2'] = u"Отчет по ребрам"
                    ws['A3'] = u"Дата и время создания отчета"
                    ws['B3'] = date_and_time_for_support                    
                    ws['A4'] = u"Количество ребер"
                    ws['B4'] = unicode(self.line_layer.featureCount())
                    ws['A5'] = u"Количество выделенных ребер"
                    ws['B5'] = unicode(self.line_layer.selectedFeatureCount())                                                                             
                          
                    workbook.save(filename)  

                if(text_type_support_item == u"Отчет по ремонтным работам" and filename != u""):                                                                  
                    Workbook().save(filename)    
                    workbook = load_workbook(filename)
                    ws = workbook.active
                                        
                    ws['B2'] = u"Отчет по ремонтным работам"
                    ws['A3'] = u"Дата и время создания отчета"
                    ws['B3'] = date_and_time_for_support                    
                    
                    ws['B5'] = u"ID"
                    ws['C5'] = u"TYPE_REMONT"
                    ws['D5'] = u"OPISANIE_REMONT"
                    ws['E5'] = u"DATE_BEGIN"
                    ws['F5'] = u"DATE_END"
                    ws['G5'] = u"OTVET_LICO"
                          
                    conn = psycopg2.connect("dbname='vodokanal_data' user='postgres' host='localhost' password='postgres'")                
                    cur_con = conn.cursor()
                    cur_con.execute("SELECT * FROM graf_remont")
                    row_remonts = cur_con.fetchall()
                    
                    
                    num_row = 6
                    print type(row_remonts)
                    for row in row_remonts:
                        ws.cell(column=2, row=num_row).value = row[0]
                        ws.cell(column=3, row=num_row).value = row[1]
                        ws.cell(column=4, row=num_row).value = row[2]
                        ws.cell(column=5, row=num_row).value = row[3]
                        ws.cell(column=6, row=num_row).value = row[4]
                        ws.cell(column=7, row=num_row).value = row[5]
                        num_row = num_row + 1
                    
                    conn.commit()
                    
                    cur_con.close()
                    conn.close()        

                          
                          
                    workbook.save(filename) 
                    
    
    #закрыть форму отчета
    def def_close_support(self):
        self.forma_create_support.hide()        

    ########################################################################################################################################################
    ##########                    упление узлами                   #########################################################################################
    ########################################################################################################################################################
        
    #добавить узел    
    def def_add_uzel(self):       
        self.panel_instrumentov.btn_regim_uzlov.setChecked(True)
        self.init_regim_uzlov()
        self.panel_instrumentov.setEnabled(False)
                
        #создать временный слой
        self.temp_point_layer = QgsVectorLayer("Point?crs=epsg:4326", "temp_vershini", "memory")   
        self.temp_point_layer.startEditing()      
        layerData = self.temp_point_layer.dataProvider()         
        layerData.addAttributes([ QgsField("id", QVariant.Int)])
        self.temp_point_layer.commitChanges()
        self.delete_temp_layer_ver = QgsMapLayerRegistry.instance().addMapLayer(self.temp_point_layer)         

        self.connect_signal_point_tool_init(self.iface.mapCanvas(), "handle_mouse_coord_ver", "get_coord_for_add_uzel", [self.plugin_dir, self.panel_instrumentov, self.delete_temp_layer_ver, self.temp_point_layer, self.point_layer, self.zdan_layer, self.iface.mainWindow()])      
      
    #определения нажатых клавиш мыши
    def get_coord_for_add_uzel(self, point, button, plugin_dir, panel_instrumentov, delete_temp_layer_ver, temp_point_layer, point_layer, zdan_layer, mainWindow, canvas):                                
        if(button == Qt.LeftButton):
            file_ui = plugin_dir + "/forms/forma_create_uzel.ui"
            self.forma_create_uzel = uic.loadUi(file_ui)
            self.forma_create_uzel.setWindowTitle(u'Добавление нового узла')        
            self.forma_create_uzel.setWindowModality(QtCore.Qt.ApplicationModal)    
             
            QtCore.QObject.connect(self.forma_create_uzel.btn_create_ver, QtCore.SIGNAL("clicked()"), lambda: self.def_create_new_uzel(point_layer, temp_point_layer, delete_temp_layer_ver, mainWindow, panel_instrumentov, zdan_layer, point, canvas))
            QtCore.QObject.connect(self.forma_create_uzel.btn_close_form, QtCore.SIGNAL("clicked()"), lambda: self.def_close_form_create_uzel(delete_temp_layer_ver, panel_instrumentov, canvas))                            

            self.forma_create_uzel.show() 
            
        if(button == Qt.RightButton):
            self.connect_signal_point_tool_init(canvas, "", "", [])      
            QgsMapLayerRegistry.instance().removeMapLayer(delete_temp_layer_ver.id())
            panel_instrumentov.setEnabled(True)
            panel_instrumentov.btn_create_uzel.setChecked(False)      
      
    #определение координат курсора
    def handle_mouse_coord_ver(self, point, temp_point_layer, canvas):        
        if(temp_point_layer.featureCount() == 0):
            temp_point = QgsPoint(point.x(), point.y())
            pt = QgsFeature()
            pt.setGeometry(QgsGeometry.fromPoint(temp_point))
            pr = temp_point_layer.dataProvider()             
            pt.setAttributes([1])
            pr.addFeatures([pt])
            temp_point_layer.updateExtents()   
          
        else:
            geom = QgsGeometry.fromPoint(QgsPoint(point.x(), point.y()))            
            temp_point_layer.dataProvider().changeGeometryValues({1: geom})             
      
        canvas.refresh()
        if hasattr(temp_point_layer, "setCacheImage"):
            temp_point_layer.setCacheImage(None)
        temp_point_layer.triggerRepaint()       
        
     
    #добавление узла на слой 
    def def_create_new_uzel(self, point_layer, temp_point_layer, delete_temp_layer_ver, mainWindow, panel_instrumentov, zdan_layer, point_cur, canvas):
        name_ver = self.forma_create_uzel.edt_name_ver.text()        
        if(name_ver != ""):
            iter = temp_point_layer.getFeatures()
            for feature in iter:
                feature_point_id = feature.id()                     
                if(feature_point_id == 1):                
                    point = QgsPoint(feature.geometry().asPoint())
     
            type_ver_in_table = unicode(self.forma_create_uzel.cmb_type_ver.currentText())                
                
            pt = QgsFeature()
            pt.setGeometry(QgsGeometry.fromPoint(point))
            pr = point_layer.dataProvider() 
                
            iter = point_layer.getFeatures()
            feature_point_max_id = 0
                
            for feature in iter:
                if(feature.id() > feature_point_max_id):
                    feature_point_max_id = feature.id()                
                
            feature_point_id = feature_point_max_id + 1
                           
            provider = zdan_layer.dataProvider()                 
            spIndex = QgsSpatialIndex()
            feat = QgsFeature()
            fit = provider.getFeatures() 
            while fit.nextFeature(feat):
                spIndex.insertFeature(feat)
            pt_cur = QgsPoint(point_cur.x(), point_cur.y())			
            nearestIds = spIndex.nearestNeighbor(pt_cur,1)            
            id_feature_cur = nearestIds[0]  
            zdania_atributes = []
            iter_cur = zdan_layer.getFeatures()
            for feature in iter_cur:                   
                if(feature.id() == id_feature_cur):                
                    zdania_atributes_ulica = feature.attributes()[3]                    
           
            pt.setAttributes([feature_point_id, name_ver, type_ver_in_table, "", zdania_atributes_ulica])
            pr.addFeatures([pt])
            point_layer.updateExtents()            
                            
            self.def_close_form_create_uzel(delete_temp_layer_ver, panel_instrumentov, canvas)  

            canvas.refresh()
            if hasattr(point_layer, "setCacheImage"):
                point_layer.setCacheImage(None)
            point_layer.triggerRepaint()   
                                 
            panel_instrumentov.setEnabled(True)
            panel_instrumentov.btn_create_uzel.setChecked(False)
                             
        else:
            QtGui.QMessageBox.warning(mainWindow, u"Некорректный ввод данных", u"Введите имя вершины!") 
     
    #закрыть форму добавления узла
    def def_close_form_create_uzel(self, delete_temp_layer_ver, panel_instrumentov, canvas):
        self.forma_create_uzel.hide()          
        self.connect_signal_point_tool_init(canvas, "", "", [])
        QgsMapLayerRegistry.instance().removeMapLayer(delete_temp_layer_ver.id())
        
        panel_instrumentov.setEnabled(True)
        panel_instrumentov.btn_regim_uzlov.setChecked(False)
        panel_instrumentov.btn_create_uzel.setChecked(False)            
    
    #----------------------------------------------------------------------------------------------------------------------------------------
    #инициализация редактирования узла   
    def def_edit_uzel(self):
        self.panel_instrumentov.btn_regim_uzlov.setChecked(True)
        self.init_regim_uzlov()
        self.panel_instrumentov.setEnabled(False)           
    
        self.connect_signal_point_tool_init(self.iface.mapCanvas(), "", "get_id_for_edit_uzel", [self.plugin_dir, self.panel_instrumentov, self.point_layer, self.zdan_layer, self.iface.mainWindow()])          
    
    
    #определить id для редактирования
    def get_id_for_edit_uzel(self, point, button, plugin_dir, panel_instrumentov, point_layer, zdan_layer, canvas):
        #определение id ближайшего элемента к курсору
        provider = point_layer.dataProvider()                 
        spIndex = QgsSpatialIndex()
        feat = QgsFeature()
        fit = provider.getFeatures() 
        while fit.nextFeature(feat):
            spIndex.insertFeature(feat)
        pt = QgsPoint(point.x(), point.y())			
        nearestIds = spIndex.nearestNeighbor(pt,1)            
        id_feature = nearestIds[0]        

        self.connect_signal_point_tool_init(canvas, "", "", [])          
        
        if(button == Qt.LeftButton):
            point_layer.removeSelection()
            point_layer.select(id_feature)
            
            iter = point_layer.getFeatures()
            for feature in iter:                   
                if(feature.id() == id_feature):                
                    self.edit_id_uzel = feature.id()  
                    self.edit_geom_uzel = feature.geometry().asPoint()
                    self.edit_feature_uzel_atr = feature.attributes()            
            
            file_ui = plugin_dir + "/forms/forma_edit_uzel.ui"            
            self.forma_edit_uzel = uic.loadUi(file_ui)
            self.forma_edit_uzel.setWindowTitle(u'Редактирование узла')        
            self.forma_edit_uzel.setWindowModality(QtCore.Qt.ApplicationModal)                       
            
            QtCore.QObject.connect(self.forma_edit_uzel.btn_set_geomery_uzel, QtCore.SIGNAL("clicked()"), lambda: self.def_edit_regim_okna(self.edit_id_uzel, self.edit_geom_uzel, self.edit_feature_uzel_atr, canvas, panel_instrumentov, point_layer, self.forma_edit_uzel))
            QtCore.QObject.connect(self.forma_edit_uzel.btn_save_ver,         QtCore.SIGNAL("clicked()"), lambda: self.def_save_edit_uzel(self.forma_edit_uzel, canvas, point_layer, panel_instrumentov))
            QtCore.QObject.connect(self.forma_edit_uzel.btn_close_form,       QtCore.SIGNAL("clicked()"), lambda: self.def_close_form_edit_uzel(self.forma_edit_uzel, panel_instrumentov)) 
            
            self.forma_edit_uzel.edt_name_ver.setText(unicode(self.edit_feature_uzel_atr[1]))            
            index = self.forma_edit_uzel.cmb_type_ver.findText(self.edit_feature_uzel_atr[2])
            self.forma_edit_uzel.cmb_type_ver.setCurrentIndex(index)            
            
            self.forma_edit_uzel.show()            
            
        if(button == Qt.RightButton):
            panel_instrumentov.setEnabled(True)
            panel_instrumentov.btn_regim_uzlov.setChecked(False)                  
            panel_instrumentov.btn_edt_uzel.setChecked(False)                  

    # активировать окно формы редактирования       
    def def_edit_regim_okna(self, feature_edit_id, feature_edit_geom, feature_edit_attribute, canvas, panel_instrumentov, point_layer, forma_edit_uzel):
        if(self.forma_edit_uzel.btn_set_geomery_uzel.isChecked()):            
            self.forma_edit_uzel.setWindowModality(QtCore.Qt.NonModal)
            self.forma_edit_uzel.setEnabled(False)
            
            self.connect_signal_point_tool_init(canvas, "def_set_geometry_uzel_coord", "def_set_geometry_uzel_click", [point_layer, feature_edit_id, feature_edit_geom, feature_edit_attribute, forma_edit_uzel, panel_instrumentov])          
                       
        else:
            self.forma_edit_uzel.activateWindow()
            self.forma_edit_uzel.setWindowModality(QtCore.Qt.ApplicationModal)
            self.forma_edit_uzel.setEnabled(True)

            self.connect_signal_point_tool_init(self.iface.mapCanvas(), "", "", [])          
             
    # изменить геометрию по координатам
    def def_set_geometry_uzel_coord(self, point, point_layer, edit_id_uzel, canvas):
            geom = QgsGeometry.fromPoint(QgsPoint(point.x(), point.y()))            
            point_layer.dataProvider().changeGeometryValues({edit_id_uzel: geom})    
            canvas.refresh()
            if hasattr(point_layer, "setCacheImage"):
                point_layer.setCacheImage(None)
            point_layer.triggerRepaint()    
    
    # зафиксировать изменение геометрии
    def def_set_geometry_uzel_click(self, point, button, point_layer, edit_id_uzel, edit_geom_uzel, forma_edit_uzel, canvas):
        forma_edit_uzel.activateWindow()
        forma_edit_uzel.setWindowModality(QtCore.Qt.ApplicationModal)
        forma_edit_uzel.setEnabled(True)
        forma_edit_uzel.btn_set_geomery_uzel.setChecked(False)

        self.connect_signal_point_tool_init(canvas, "", "", [])        
        
        if(button == Qt.LeftButton):
            geom = QgsGeometry.fromPoint(QgsPoint(point.x(), point.y()))            
            point_layer.dataProvider().changeGeometryValues({edit_id_uzel: geom})  
               
        if(button == Qt.RightButton):    
            geom_s = QgsGeometry.fromPoint(QgsPoint(edit_geom_uzel.x(), edit_geom_uzel.y()))            
            point_layer.dataProvider().changeGeometryValues({edit_id_uzel: geom_s})            
  
    # закрыть окно редактирования узла
    def def_close_form_edit_uzel(self, forma_edit_uzel, panel_instrumentov):
        if(not(forma_edit_uzel.btn_set_geomery_uzel.isChecked())):
            panel_instrumentov.setEnabled(True)
            panel_instrumentov.btn_edt_uzel.setChecked(False) 
            forma_edit_uzel.hide()
  
    # сохранить изменения редактирования узла
    def def_save_edit_uzel(self, forma_edit_uzel, canvas, point_layer, panel_instrumentov):
        name_ver = forma_edit_uzel.edt_name_ver.text()        
        if(name_ver != ""):                
            type_ver_in_table = unicode(forma_edit_uzel.cmb_type_ver.currentText())   
            attrs = { 1 : name_ver, 2 : type_ver_in_table }
            point_layer.dataProvider().changeAttributeValues({ self.edit_id_uzel : attrs })    
            point_layer.updateExtents() 
                
            canvas.refresh()
            if hasattr(point_layer, "setCacheImage"):
                point_layer.setCacheImage(None)
            point_layer.triggerRepaint()

            self.connect_signal_point_tool_init(canvas, "", "", [])
            
            panel_instrumentov.setEnabled(True)
            panel_instrumentov.btn_edt_uzel.setChecked(False) 
            forma_edit_uzel.hide()
      
        else:
            QtGui.QMessageBox.information(self.iface.mainWindow(), u"Некорректный ввод данных", u"Введите имя вершины!")        
          

    def def_remont_uzla(self):
        self.panel_instrumentov.setEnabled(False)           
        self.connect_signal_point_tool_init(self.iface.mapCanvas(), "", "get_id_for_rem_uzel", [self.plugin_dir, self.panel_instrumentov, self.point_layer, self.zdan_layer, self.iface.mainWindow()])          

    def get_id_for_rem_uzel(self, point, button, plugin_dir, panel_instrumentov, point_layer, zdan_layer, mainWindow, canvas):
        #определение id ближайшего элемента к курсору
        provider = point_layer.dataProvider()                 
        spIndex = QgsSpatialIndex()
        feat = QgsFeature()
        fit = provider.getFeatures() 
        while fit.nextFeature(feat):
            spIndex.insertFeature(feat)
        pt = QgsPoint(point.x(), point.y())			
        nearestIds = spIndex.nearestNeighbor(pt,1)            
        id_feature = nearestIds[0]                       
        
        if(button == Qt.LeftButton):
            point_layer.removeSelection()
            point_layer.select(id_feature)
            
            iter = point_layer.getFeatures()
            for feature in iter:                   
                if(feature.id() == id_feature):                
                    self.edit_id_uzel = feature.id()  
                    self.edit_geom_uzel = feature.geometry().asPoint()
                    self.edit_feature_uzel_atr = feature.attributes()            
            
            file_ui = plugin_dir + "/forms/forma_ucheta_remontov.ui"            
            self.forma_ucheta_remontov = uic.loadUi(file_ui)
            self.forma_ucheta_remontov.setWindowTitle(u'Добавление ремонта')        
            self.forma_ucheta_remontov.setWindowModality(QtCore.Qt.ApplicationModal)                       
            
            QtCore.QObject.connect(self.forma_ucheta_remontov.btn_create_support, QtCore.SIGNAL("clicked()"), lambda: self.def_add_remont(self.forma_ucheta_remontov, mainWindow, id_feature, panel_instrumentov, canvas))
            QtCore.QObject.connect(self.forma_ucheta_remontov.btn_cancel_support, QtCore.SIGNAL("clicked()"), lambda: self.def_close_remont(self.forma_ucheta_remontov, panel_instrumentov, canvas))
            
            self.forma_ucheta_remontov.show()            
            
        if(button == Qt.RightButton):
            panel_instrumentov.setEnabled(True)
            panel_instrumentov.btn_rem_uzel.setChecked(False)
            self.connect_signal_point_tool_init(canvas, "", "", [])  

    def def_add_remont(self, forma_ucheta_remontov, mainWindow, id_feature, panel_instrumentov, canvas):
        opisanie = unicode(forma_ucheta_remontov.txt_opis_rem.toPlainText())
        otvetstvennoe_lico = unicode(forma_ucheta_remontov.edt_otvet_lico.text())
        type_remont = unicode(forma_ucheta_remontov.cmb_type_rem.currentText())                
        
        temp_var = forma_ucheta_remontov.date_start.date() 
        date_start = unicode(temp_var.toPyDate())
        
        temp_var = forma_ucheta_remontov.date_end.date() 
        date_end = unicode(temp_var.toPyDate())
        
        id_feature
        
        if(opisanie != u"" and otvetstvennoe_lico != u""):
            conn = psycopg2.connect("dbname='vodokanal_data' user='postgres' host='localhost' password='postgres'")                
            cur_con = conn.cursor()
            cur_con.execute("SELECT * FROM graf_remont")
            numrows_con = unicode(int(cur_con.rowcount) + 1)
             
            sql = "INSERT INTO graf_remont (\"id\", \"TYPE_REMONT\", \"OPISANIE_REMONT\", \"DATE_BEGIN\", \"DATE_END\", \"OTVET_LICO\", \"id_uzla\") VALUES (" +numrows_con+ ", '" +type_remont+ "', '"+opisanie+"', '"+date_start+"', '"+date_end+"', '"+otvetstvennoe_lico+"', "+unicode(id_feature)+ ")"
            
            cur_con.execute(sql)  
            conn.commit()
            
            cur_con.close()
            conn.close()        

            self.def_close_remont(forma_ucheta_remontov, panel_instrumentov, canvas)
        else:
            QtGui.QMessageBox.information(mainWindow, u"Ошибка ввода", u"Данные введены не корректно!") 
          
    
    def def_close_remont(self, forma_ucheta_remontov, panel_instrumentov, canvas):
        panel_instrumentov.setEnabled(True)
        panel_instrumentov.btn_rem_uzel.setChecked(False)
        forma_ucheta_remontov.hide()
        self.connect_signal_point_tool_init(canvas, "", "", [])
    
    
    # удаление узла
    def def_del_uzel(self):
        self.panel_instrumentov.btn_regim_uzlov.setChecked(False)
        self.init_regim_uzlov()
        self.panel_instrumentov.setEnabled(False)           
        
        self.connect_signal_point_tool_init(self.iface.mapCanvas(), "", "get_id_for_del_uzel", [self.point_layer, self.iface.mainWindow(), self.panel_instrumentov])        
        
    def get_id_for_del_uzel(self, point, button, point_layer, mainWindow, panel_instrumentov, canvas):
        provider = point_layer.dataProvider()                 
        spIndex = QgsSpatialIndex()
        feat = QgsFeature()
        fit = provider.getFeatures() 
        while fit.nextFeature(feat):
            spIndex.insertFeature(feat)
        pt = QgsPoint(point.x(), point.y())			
        nearestIds = spIndex.nearestNeighbor(pt,1)            
        id_feature = nearestIds[0]        
 
        if(button == Qt.LeftButton):
            point_layer.removeSelection()
            point_layer.select(id_feature)
            
            quit_msg = u"Удалить узел с ID - " +  unicode(id_feature) + u" ?"
            reply = QtGui.QMessageBox.question(mainWindow, u'Удаление узла',  quit_msg, QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)

            if reply == QtGui.QMessageBox.Yes:
                iter = point_layer.getFeatures()
                for feature in iter:                   
                    if(feature.id() == id_feature):                
                        edit_id_uzel = feature.id()
                    
                point_layer.dataProvider().deleteFeatures([edit_id_uzel])
                point_layer.updateExtents()
            
                canvas.refresh()
                if hasattr(point_layer, "setCacheImage"):
                    point_layer.setCacheImage(None)
                point_layer.triggerRepaint()             
                
            panel_instrumentov.setEnabled(True)
            panel_instrumentov.btn_del_uzel.setChecked(False)     
            
            self.connect_signal_point_tool_init(canvas, "", "", [])
            
        if(button == Qt.RightButton):    
            panel_instrumentov.setEnabled(True)
            panel_instrumentov.btn_del_uzel.setChecked(False)     
            
            self.connect_signal_point_tool_init(canvas, "", "", [])
            
    # копирование узла
    def def_copy_uzel(self):
        self.panel_instrumentov.btn_regim_uzlov.setChecked(True)
        self.init_regim_uzlov()
        self.panel_instrumentov.setEnabled(False)           
  
        self.connect_signal_point_tool_init(self.iface.mapCanvas(), "", "get_id_for_copy_uzel", [self.point_layer, self.iface.mainWindow(), self.panel_instrumentov])
        
    def get_id_for_copy_uzel(self, point, button, point_layer, mainWindow, panel_instrumentov, canvas):
        provider = point_layer.dataProvider()                 
        spIndex = QgsSpatialIndex()
        feat = QgsFeature()
        fit = provider.getFeatures() 
        while fit.nextFeature(feat):
            spIndex.insertFeature(feat)
        pt = QgsPoint(point.x(), point.y())			
        nearestIds = spIndex.nearestNeighbor(pt,1)            
        self.id_uzel_for_copy = nearestIds[0]        
 
        if(button == Qt.LeftButton):
            point_layer.removeSelection()
            point_layer.select(self.id_uzel_for_copy)
            
            quit_msg = u"Копировать узел с ID - " +  unicode(self.id_uzel_for_copy) + u" ?"
            reply = QtGui.QMessageBox.question(mainWindow, u'Копирование узла',  quit_msg, QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)

            if reply == QtGui.QMessageBox.Yes:
                iter = point_layer.getFeatures()
                for feature in iter:
                    feature_point_id = feature.id()                     
                    if(feature_point_id == self.id_uzel_for_copy):                
                        self.point_copy = QgsPoint(feature.geometry().asPoint())
                        edit_feature_uzel_atr = feature.attributes()
                        break
         
                name_ver = unicode(edit_feature_uzel_atr[1])                
                type_ver_in_table = unicode(edit_feature_uzel_atr[2])                
                    
                pt = QgsFeature()
                pt.setGeometry(QgsGeometry.fromPoint(self.point_copy))
                pr = point_layer.dataProvider() 
                    
                iter = point_layer.getFeatures()
                feature_point_max_id = 0
                    
                for feature in iter:
                    if(feature.id() > feature_point_max_id):
                        feature_point_max_id = feature.id()                
                    
                self.feature_point_id_max_for_copy = feature_point_max_id + 1
                pt.setAttributes([self.feature_point_id_max_for_copy, name_ver, type_ver_in_table])
                pr.addFeatures([pt])
                point_layer.updateExtents()

                self.connect_signal_point_tool_init(canvas, "handle_mouse_coord_ver_for_copy", "get_coord_for_add_uzel_for_copy", [point_layer, panel_instrumentov, self.feature_point_id_max_for_copy])
            
                canvas.refresh()
                if hasattr(point_layer, "setCacheImage"):
                    point_layer.setCacheImage(None)
                point_layer.triggerRepaint()
            else:
                self.connect_signal_point_tool_init(canvas, "", "", [])    
                
            panel_instrumentov.setEnabled(True)
            panel_instrumentov.btn_copy_uzel.setChecked(False)                 
        
        if(button == Qt.RightButton):    
            panel_instrumentov.setEnabled(True)
            panel_instrumentov.btn_copy_uzel.setChecked(False)     
            
            self.connect_signal_point_tool_init(canvas, "", "", []) 
            
    #определение координат курсора для копирования
    def handle_mouse_coord_ver_for_copy(self, point, point_layer, feature_point_id_max_for_copy, canvas):
        geom = QgsGeometry.fromPoint(QgsPoint(point.x(), point.y()))            
        point_layer.dataProvider().changeGeometryValues({feature_point_id_max_for_copy: geom})             
      
        canvas.refresh()
        if hasattr(point_layer, "setCacheImage"):
            point_layer.setCacheImage(None)
        point_layer.triggerRepaint()
        
    #определения нажатых клавиш мыши для копирования
    def get_coord_for_add_uzel_for_copy(self, point, button, panel_instrumentov, canvas):                                
        self.connect_signal_point_tool_init(canvas, "", "", [])
        panel_instrumentov.setEnabled(True)
        panel_instrumentov.btn_copy_uzel.setChecked(False)
        
    ########################################################################################################################################################
    ##########                    управление ребрами                   #####################################################################################
    ########################################################################################################################################################        

    #добавить новое ребро
    def def_add_reb(self):             
        self.panel_instrumentov.setEnabled(False)
        self.panel_instrumentov.btn_create_reb.setChecked(True)  

        #создать временный слой
        self.temp_point_layer_for_reb = QgsVectorLayer("Point?crs=epsg:4326", "temp_vershini_for_reb", "memory")   
        self.temp_point_layer_for_reb.startEditing()      
        layerData = self.temp_point_layer_for_reb.dataProvider()         
        layerData.addAttributes([ QgsField("id", QVariant.Int)])
        self.temp_point_layer_for_reb.commitChanges()
        self.delete_temp_point_layer_for_reb = QgsMapLayerRegistry.instance().addMapLayer(self.temp_point_layer_for_reb) 
            
        self.temp_line_layer = QgsVectorLayer("LineString?crs=epsg:4326", "temp_rebra", "memory")   
        self.temp_line_layer.startEditing()  
        layerData = self.temp_line_layer.dataProvider() 
        layerData.addAttributes([QgsField("id", QVariant.Int)])
        self.temp_line_layer.commitChanges()
        self.delete_temp_line_layer_for_reb = QgsMapLayerRegistry.instance().addMapLayer(self.temp_line_layer)        
        
        flag_point_start = False
        flag_point_finish = False
        id_point_start = 0
        id_point_finish = 0
        
        self.connect_signal_point_tool_init(self.iface.mapCanvas(), "", "get_id_for_add_reb", [self.point_layer, self.line_layer, self.temp_point_layer_for_reb, self.temp_line_layer, self.delete_temp_point_layer_for_reb, self.delete_temp_line_layer_for_reb, self.panel_instrumentov, self.iface.mainWindow(), flag_point_start, flag_point_finish, id_point_start, id_point_finish])        
        
    def get_id_for_add_reb(self, point, button, point_layer, line_layer, temp_point_layer, temp_line_layer, delete_temp_point_layer, delete_temp_line_layer, panel_instrumentov, mainWindow, flag_point_start, flag_point_finish, id_point_start, id_point_finish, canvas):       
        point_layer.removeSelection()
        
        if(point_layer.featureCount() > 0):
            provider = point_layer.dataProvider()                 
            spIndex = QgsSpatialIndex()
            feat = QgsFeature()
            fit = provider.getFeatures() 
            while fit.nextFeature(feat):
                spIndex.insertFeature(feat)
            pt = QgsPoint(point.x(), point.y())			
            nearestIds = spIndex.nearestNeighbor(pt,1)            
            point_id_feature = nearestIds[0]        
            
            if(button == Qt.LeftButton):                                       
                iter = point_layer.getFeatures()
                for feature in iter:                     
                    if(feature.id() == point_id_feature):                
                        point_geom = QgsPoint(feature.geometry().asPoint())                    
                        break
                
                dlina = sqrt( ((point.x()-point_geom.x())**2 + (point.y()-point_geom.y())**2) * canvas.scale() )            
                point_dlina_milimetr = dlina * canvas.scale() * 100
                            
                print "-------------------" 
                print dlina
                print point_dlina_milimetr                                        
                                    
                if(flag_point_start == True):                                   
                    point_layer.select(id_point_start)                                                            
                    
                    if(point_dlina_milimetr < 1 and point_id_feature != id_point_start):
                        point_layer.select(point_id_feature)                                                         
                        
                        canvas.refresh()
                        if hasattr(temp_line_layer, "setCacheImage"):
                            temp_line_layer.setCacheImage(None)
                        temp_line_layer.triggerRepaint()
                        
                        iter = temp_point_layer.getFeatures()
                        max_id_uzel = 0                                                              
                        for feature in iter:
                            if(feature.id() > max_id_uzel):
                                max_id_uzel = feature.id()
                        
                        provider = temp_point_layer.dataProvider()                   
                        temp_point_layer.dataProvider().deleteFeatures([max_id_uzel])
                        
                        iter = point_layer.getFeatures()
                        for feature in iter:                   
                            if(feature.id() == id_point_start):                                                
                                point_out = feature.geometry().asPoint()                        
                            if(feature.id() == point_id_feature):                                                
                                point_in = feature.geometry().asPoint()                        
                        
                        
                        iter = temp_point_layer.getFeatures()
                        list_temp_ver_for_reb = []
                        list_temp_ver_for_reb.append(point_out)
                        for feature in iter:
                            list_temp_ver_for_reb.append(QgsPoint(feature.geometry().asPoint()))
                        
                        list_temp_ver_for_reb.append(point_in) 
                        
                        geom = QgsGeometry.fromPolyline(list_temp_ver_for_reb)
                        temp_line_layer.dataProvider().changeGeometryValues({1:geom})
                        
                        iter = temp_line_layer.getFeatures()
                        for feature in iter:
                            feature_point_id = feature.id()                     
                            if(feature_point_id == 1):                
                                new_line_geom = feature.geometry().asPolyline()
                        try:
                            line = QgsFeature()
                            line.setGeometry(QgsGeometry.fromPolyline(new_line_geom))

                            
                            iter = line_layer.getFeatures()
                                
                            feature_point_max_id = 0
                            for feature in iter:
                                if(feature.id() > feature_point_max_id):
                                    feature_point_max_id = feature.id()                                
                            feature_point_id = feature_point_max_id + 1                                                       

                            ver_out_id = id_point_start
                            ver_in_id = point_id_feature
                            
                            iter = point_layer.getFeatures()
                            for feature in iter:                    
                                if(feature.id() == id_point_start):                
                                    feature_point_atr = feature.attributes()                
                                       
                            type_ver = unicode(feature_point_atr[2])
                            smeg_ver = feature_point_atr[3] 
                            
                            if(smeg_ver == None):
                                count_smeg = 0
                            else:            
                                list_smeg = smeg_ver.split("/")
                                count_smeg = len(list_smeg)

                            if(count_smeg == 0):
                                smeg_ver = ver_in_id                
                            else:
                                smeg_ver = unicode(smeg_ver) + u"/" + unicode(ver_in_id)
                  
                            attrs = { 3 : smeg_ver }
                            point_layer.dataProvider().changeAttributeValues({ ver_out_id : attrs })    
                            point_layer.updateExtents()
                            
                            type_reb = None
                            
                            if((type_ver == u"Канализация") or (type_ver == u"Потребитель")):
                                type_reb = u"Отвод:течет"
                           
                            if((type_ver == u"Колодец") or (type_ver == u"Источник") or (type_ver == u"Развилка") or (type_ver == u"Насосная станция")):
                                type_reb = u"Подвод:течет"           
                           
                            line.setAttributes([feature_point_id, "", ver_out_id, ver_in_id, type_reb])
                            pr = line_layer.dataProvider() 
                            pr.addFeatures([line])
                            
                            line_layer.updateExtents() 
                                  
                            line_layer.dataProvider().changeGeometryValues({feature_point_id: geom})
                        except:
                            print "error: get_id_for_add_reb"
                            
                        canvas.refresh()
                        if hasattr(line_layer, "setCacheImage"):
                            line_layer.setCacheImage(None)
                        line_layer.triggerRepaint()
                        
                        
                        panel_instrumentov.setEnabled(True)
                        panel_instrumentov.btn_create_reb.setChecked(False)
                        self.connect_signal_point_tool_init(canvas, "", "", [])
                        
                        QgsMapLayerRegistry.instance().removeMapLayer(delete_temp_point_layer.id())
                        QgsMapLayerRegistry.instance().removeMapLayer(delete_temp_line_layer.id())
                        
                        
                    else:
                        iter = temp_point_layer.getFeatures()
                        id_new_temp_point = 0                                
                        for feature in iter:
                            if(feature.id() > id_new_temp_point):
                                id_new_temp_point = feature.id() 
                          
                        id_new_temp_point = id_new_temp_point + 1  
                        temp_point = QgsPoint(point.x(), point.y())
                        pt = QgsFeature()
                        pt.setGeometry(QgsGeometry.fromPoint(temp_point))
                        pr = temp_point_layer.dataProvider()             
                        pt.setAttributes([id_new_temp_point])
                        pr.addFeatures([pt])
                        temp_point_layer.updateExtents()                                                      
                        
                        #обновление канвы
                        canvas.refresh()
                        if hasattr(temp_point_layer, "setCacheImage"):
                            temp_point_layer.setCacheImage(None)
                        temp_point_layer.triggerRepaint()  

                        canvas.refresh()
                        if hasattr(temp_line_layer, "setCacheImage"):
                            temp_line_layer.setCacheImage(None)
                        temp_line_layer.triggerRepaint()                          
                    
                else:
                    if(point_dlina_milimetr < 1):               
                        flag_point_start = True
                        point_layer.select(point_id_feature)
                        self.connect_signal_point_tool_init(canvas, "get_coord_for_add_reb", "get_id_for_add_reb", [point_layer, line_layer, temp_point_layer, temp_line_layer, delete_temp_point_layer, delete_temp_line_layer, panel_instrumentov, mainWindow, flag_point_start, flag_point_finish, point_id_feature, id_point_finish])                                                         
                      
                        canvas.refresh()
                        if hasattr(temp_line_layer, "setCacheImage"):
                            temp_line_layer.setCacheImage(None)
                        temp_line_layer.triggerRepaint()      
            
            if(button == Qt.RightButton):
                if(temp_point_layer.featureCount() > 0):
                    iter = temp_point_layer.getFeatures()
                    max_id_uzel = 0                                
                    predmax_id_uzel = 0                                
                    for feature in iter:
                        if(feature.id() > max_id_uzel):
                            predmax_id_uzel = max_id_uzel
                            max_id_uzel = feature.id()
                    
                    provider = temp_point_layer.dataProvider()                   
                    temp_point_layer.dataProvider().deleteFeatures([predmax_id_uzel])

                    #обновление канвы
                    canvas.refresh()
                    if hasattr(temp_point_layer, "setCacheImage"):
                        temp_point_layer.setCacheImage(None)
                    temp_point_layer.triggerRepaint()    

                    canvas.refresh()
                    if hasattr(temp_line_layer, "setCacheImage"):
                        temp_line_layer.setCacheImage(None)
                    temp_line_layer.triggerRepaint() 
                    canvas.refresh()
        else:
            QtGui.QMessageBox.information(self.iface.mainWindow(), u"Добавление нового ребра", u"Невозможно добавить ребро, слой узлов пуст!")        
    
    
    def get_coord_for_add_reb(self, point, point_layer, line_layer, temp_point_layer, temp_line_layer, delete_temp_point_layer, delete_temp_line_layer, panel_instrumentov, mainWindow, flag_point_start, flag_point_finish, id_point_start, id_point_finish, canvas):       
        if(flag_point_start == True and temp_point_layer.featureCount() > 0):        
            iter = point_layer.getFeatures()
            for feature in iter:
                feature_point_id = feature.id()             
                if(feature_point_id == id_point_start):                
                    point_out = QgsPoint(feature.geometry().asPoint())            
                if(feature_point_id == id_point_finish):                
                    point_in = QgsPoint(feature.geometry().asPoint())                        
             
            iter = temp_point_layer.getFeatures()
            edit_id_uzel = 0                                
            for feature in iter:
                if(feature.id() > edit_id_uzel):
                    edit_id_uzel = feature.id() 
            
            geom = QgsGeometry.fromPoint(QgsPoint(point.x(), point.y()))            
            temp_point_layer.dataProvider().changeGeometryValues({edit_id_uzel: geom})                            
                        

            #обновление канвы
            canvas.refresh()
            if hasattr(temp_point_layer, "setCacheImage"):
                temp_point_layer.setCacheImage(None)
            temp_point_layer.triggerRepaint()
            

            iter = temp_point_layer.getFeatures()
            list_temp_ver_for_reb = []
            list_temp_ver_for_reb.append(point_out)
            for feature in iter:
                list_temp_ver_for_reb.append(QgsPoint(feature.geometry().asPoint()))
            
            if(flag_point_finish == True):        
                list_temp_ver_for_reb.append(point_in) 

            if(temp_line_layer.featureCount() == 0):
                pr = temp_line_layer.dataProvider()                
                fet = QgsFeature()
                fet.setGeometry( QgsGeometry.fromPolyline( list_temp_ver_for_reb ))
                pr.addFeatures( [ fet ] )
            else:    
                geom = QgsGeometry.fromPolyline(list_temp_ver_for_reb)            
                temp_line_layer.dataProvider().changeGeometryValues({1: geom })
                
            canvas.refresh()
            if hasattr(temp_line_layer, "setCacheImage"):
                temp_line_layer.setCacheImage(None)
            temp_line_layer.triggerRepaint()        
    
    # удаление ребра    
    def def_del_reb(self):
        self.panel_instrumentov.btn_regim_uzlov.setChecked(False)
        self.init_regim_uzlov()
        self.panel_instrumentov.setEnabled(False)           
        
        self.connect_signal_point_tool_init(self.iface.mapCanvas(), "", "get_id_for_del_rebro", [self.line_layer, self.iface.mainWindow(), self.panel_instrumentov])        
        
    def get_id_for_del_rebro(self, point, button, line_layer, mainWindow, panel_instrumentov, canvas):
        provider = line_layer.dataProvider()                 
        spIndex = QgsSpatialIndex()
        feat = QgsFeature()
        fit = provider.getFeatures() 
        while fit.nextFeature(feat):
            spIndex.insertFeature(feat)
        pt = QgsPoint(point.x(), point.y())			
        nearestIds = spIndex.nearestNeighbor(pt,1)            
        id_feature = nearestIds[0]        
 
        if(button == Qt.LeftButton):
            line_layer.removeSelection()
            line_layer.select(id_feature)
            
            quit_msg = u"Удалить ребро с ID - " +  unicode(id_feature) + u" ?"
            reply = QtGui.QMessageBox.question(mainWindow, u'Удаление ребра',  quit_msg, QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)

            if reply == QtGui.QMessageBox.Yes:
                iter = line_layer.getFeatures()
                for feature in iter:                   
                    if(feature.id() == id_feature):                
                        edit_id_reb = feature.id()
                    
                line_layer.dataProvider().deleteFeatures([edit_id_reb])
                line_layer.updateExtents()
            
                canvas.refresh()
                if hasattr(line_layer, "setCacheImage"):
                    line_layer.setCacheImage(None)
                line_layer.triggerRepaint()             
                
            panel_instrumentov.setEnabled(True)
            panel_instrumentov.btn_del_reb.setChecked(False)     
            
            self.connect_signal_point_tool_init(canvas, "", "", [])
            
        if(button == Qt.RightButton):    
            panel_instrumentov.setEnabled(True)
            panel_instrumentov.btn_del_uzel.setChecked(False)     
            
            self.connect_signal_point_tool_init(canvas, "", "", [])